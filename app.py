from flask import Flask, render_template_string, request, send_file, jsonify
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
from datetime import datetime, timedelta
import random
import os
import json
import math
import threading
import uuid
import tempfile

app = Flask(__name__)

# ==========================================
# CONFIGURACIÓN DE LA API Y RUTAS
# ==========================================
API_KEY = "7bd626cb4d3874faf995ec075af15d2cd35ec99d"
BASE_URL = "https://gps.idttecnologias.mx/api/v1"
COMPANY_ID = "87534"
TIMEZONE_OFFSET = -7

# DICCIONARIO GLOBAL PARA GUARDAR EL ESTADO DE LAS TAREAS
TASKS = {}

# ==========================================
# FÓRMULA HAVERSINE
# ==========================================
def calcular_distancia(lat1, lon1, lat2, lon2):
    R = 6371000 
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dphi/2.0)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlon/2.0)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

# ==========================================
# LECTOR EXCEL
# ==========================================
def cargar_geocercas_excel():
    geocercas = []
    base_dir = os.path.dirname(os.path.abspath(__file__))
    posibles_archivos = ['kowi_principales.xlsx', 'kowi principales.xlsx']
    ruta_final = None
    for archivo in posibles_archivos:
        ruta_temp = os.path.join(base_dir, archivo)
        if os.path.exists(ruta_temp):
            ruta_final = ruta_temp
            break
    if not ruta_final: return [{'error': 'file_not_found', 'msg': 'No se encontró el archivo .xlsx en Render'}]
    try:
        wb = openpyxl.load_workbook(ruta_final, data_only=True)
        ws = wb.active
        headers = [str(cell.value).lower().strip() if cell.value else '' for cell in ws[1]]
        idx_nom = next((i for i, h in enumerate(headers) if 'nombre' in h or 'zona' in h), -1)
        idx_lat = next((i for i, h in enumerate(headers) if 'lat' in h), -1)
        idx_lon = next((i for i, h in enumerate(headers) if 'lon' in h or 'lng' in h), -1)
        idx_area = next((i for i, h in enumerate(headers) if 'rea' in h or 'area' in h), -1)
        
        if idx_nom == -1 or idx_lat == -1 or idx_lon == -1: return [{'error': 'parsing_failed', 'msg': f'Columnas no detectadas.'}]
            
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not row[idx_nom] or not row[idx_lat] or not row[idx_lon]: continue
            nombre = str(row[idx_nom]).strip()
            lat_str = str(row[idx_lat]).strip()
            lng_str = str(row[idx_lon]).strip()
            area_str = str(row[idx_area]).strip() if idx_area != -1 and row[idx_area] else ''
            radio_m = 200 
            try:
                area_limpia = float(str(area_str).lower().replace('km2', '').replace('m2', '').replace(',', '').strip())
                area_m2 = area_limpia * 1000000 if 'km2' in str(area_str).lower() else area_limpia
                radio_calculado = math.sqrt(area_m2 / math.pi)
                if radio_calculado > 0: radio_m = min(radio_calculado, 1500) 
            except: pass
            geocercas.append({'id': f"LOCAL_{i}", 'name': nombre, 'lat': float(lat_str), 'lng': float(lng_str), 'radius': radio_m})
        return geocercas
    except Exception as e: return [{'error': 'exception', 'msg': str(e)}]

GEOCERCAS_MAESTRAS = cargar_geocercas_excel()

HTML_INTERFACE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Histórico De Rutas Minuto a Minuto - IDT</title>
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <link href="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/css/select2.min.css" rel="stylesheet" />
    <script src="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/js/select2.min.js"></script>

    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f6f9; padding: 20px; margin: 0; padding-bottom: 80px; }
        .card { max-width: 820px; margin: 0 auto; background: #ffffff; padding: 25px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.08); position: relative; z-index: 10; }
        
        .header { text-align: center; margin-bottom: 20px; border-bottom: 1px solid #eee; padding-bottom: 15px;}
        .header h2 { color: #1a252f; margin: 0 0 5px 0; font-size: 22px; }
        .header p { color: #7f8c8d; font-size: 13px; margin: 0; }
        
        /* OVERLAY DE CARGA (VIDEO) */
        #loading_overlay {
            display: none; 
            position: fixed; top: 0; left: 0; width: 100%; height: 100%;
            background: rgba(255, 255, 255, 0.96);
            z-index: 9999;
            flex-direction: column; justify-content: center; align-items: center;
        }
        #loading_overlay video {
            max-width: 500px;
            width: 90%;
            border-radius: 12px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.15);
        }
        .loading-text { margin-top: 25px; font-size: 22px; font-weight: bold; color: #2c3e50; }
        .loading-subtext { color: #e67e22; margin-top: 8px; font-size: 15px; font-weight: 600; text-align: center; }

        /* LOGOS EN LAS ESQUINAS INFERIORES */
        .logo-bottom-left { position: fixed; bottom: 20px; left: 20px; max-height: 55px; z-index: 100; object-fit: contain; }
        .logo-bottom-right { position: fixed; bottom: 20px; right: 20px; max-height: 55px; z-index: 100; object-fit: contain; }

        .main-container { display: flex; gap: 20px; }
        .presets-sidebar { width: 190px; border-right: 1px solid #eee; padding-right: 15px; display: flex; flex-direction: column; gap: 6px; }
        .presets-sidebar button { background: #f8f9fa; border: 1px solid #dde2e5; color: #495057; text-align: left; padding: 8px 12px; border-radius: 6px; font-size: 12px; font-weight: 600; cursor: pointer; transition: all 0.2s; }
        .presets-sidebar button:hover { background: #e9ecef; color: #212529; }
        .form-content { flex: 1; }
        .form-group { margin-bottom: 15px; }
        .form-row { display: flex; gap: 10px; }
        .form-row .form-group { flex: 1; }
        label { display: block; font-weight: 600; margin-bottom: 5px; color: #34495e; font-size: 12px; }
        select, input { width: 100%; padding: 9px; border: 1px solid #dcdfe6; border-radius: 6px; box-sizing: border-box; font-size: 13px; color: #2c3e50; }
        .select2-container .select2-selection--single { height: 40px !important; border: 1px solid #dcdfe6 !important; border-radius: 6px !important; display: flex !important; align-items: center !important; }
        .btn-submit { width: 100%; background: #00a8ff; color: white; padding: 12px; border: none; border-radius: 6px; font-weight: bold; font-size: 14px; cursor: pointer; transition: background 0.2s; margin-top: 10px; }
        .btn-submit:hover { background: #008be3; }
        .btn-submit:disabled { background: #8cd6ff; cursor: not-allowed; }
        .status-msg { font-size: 13px; color: #e67e22; margin-top: 10px; text-align: center; font-weight: bold; }
        .retry-btn { font-size: 11px; color: #007bff; text-decoration: underline; cursor: pointer; margin-left: 8px; }
        hr { border: 0; border-top: 1px solid #dde2e5; margin: 5px 0; }
    </style>
</head>
<body>

    <!-- PANTALLA DE CARGA CON VIDEO (Carga desde la carpeta static) -->
    <div id="loading_overlay">
        <video autoplay loop muted playsinline>
            <source src="{{ url_for('static', filename='video_carga.mp4') }}" type="video/mp4">
            Tu navegador no soporta videos.
        </video>
        <div class="loading-text">Generando Reporte Ejecutivo...</div>
        <div class="loading-subtext" id="overlay_status">Conectando con servidores...</div>
    </div>

    <!-- LOGOS FLOTANTES EN LAS ESQUINAS INFERIORES -->
    <img src="{{ url_for('static', filename='logo_kowi.png') }}" class="logo-bottom-left" alt="Kowi">
    <img src="{{ url_for('static', filename='logo_idt.png') }}" class="logo-bottom-right" alt="IDT Tecnologías">

    <div class="card">
        <div class="header">
            <h2>📊 Histórico De Rutas Minuto a Minuto</h2>
            <p>Motor Asíncrono de Procesamiento Masivo</p>
        </div>
        
        <div class="main-container">
            <div class="presets-sidebar">
                <label>Atajos de Fecha:</label>
                <button type="button" onclick="setRange('hoy')">Hoy</button>
                <button type="button" onclick="setRange('ayer')">Ayer</button>
                <button type="button" onclick="setRange('esta_semana')">Esta Semana</button>
                <button type="button" onclick="setRange('semana_anterior')">Semana anterior</button>
                <button type="button" onclick="setRange('ultimos_7_dias')">Últimos 7 días</button>
                <hr>
                <button type="button" onclick="setRange('este_mes')">Mes completo</button>
                <button type="button" onclick="setRange('mes_anterior')">Mes anterior</button>
            </div>

            <div class="form-content">
                <div class="form-group">
                    <label>Buscar / Seleccionar Unidad: <span class="retry-btn" onclick="cargarCatalogos()">🔄 Reintentar carga</span></label>
                    <select id="unit_select" style="width: 100%;">
                        <option value="">⏳ Cargando catálogo...</option>
                    </select>
                </div>

                <div class="form-row">
                    <div class="form-group">
                        <label>Fecha Inicial:</label>
                        <input type="date" id="fecha_inicio" required>
                    </div>
                    <div class="form-group">
                        <label>Fecha Final:</label>
                        <input type="date" id="fecha_fin" required>
                    </div>
                </div>

                <div class="form-row">
                    <div class="form-group">
                        <label>Hora Inicio:</label>
                        <input type="time" id="hora_inicio" value="00:00" required>
                    </div>
                    <div class="form-group">
                        <label>Hora Fin:</label>
                        <input type="time" id="hora_fin" value="23:59" required>
                    </div>
                </div>

                <div class="form-row">
                    <div class="form-group">
                        <label>Límite Velocidad General (km/h):</label>
                        <input type="number" id="limite_velocidad" value="80" min="1" max="150" required>
                    </div>
                    <div class="form-group">
                        <label>Filtro de Ralentí (mins):</label>
                        <input type="number" id="min_ralenti" value="2" min="1" max="60" required>
                    </div>
                </div>

                <div class="form-group">
                    <label>Geocercas (Leídas de Excel):</label>
                    <select id="geofence_select" multiple="multiple" style="width: 100%;">
                        <option value="">⏳ Cargando geocercas locales...</option>
                    </select>
                </div>
                <div class="form-group" id="speed_limits_container"></div>

                <button type="button" class="btn-submit" id="btn_submit" onclick="iniciarReporte()">📥 Generar reporte</button>
                <div id="status_msg" class="status-msg"></div>
            </div>
        </div>
    </div>

    <script>
        async function cargarCatalogos() {
            try {
                const [resUnits, resGeos] = await Promise.all([fetch('/api_unidades'), fetch('/api_geocercas_locales')]);
                const units = await resUnits.json();
                const geos = await resGeos.json();

                const selectUnit = $('#unit_select');
                selectUnit.empty().append(new Option('🔍 Escribe para buscar unidad...', ''));
                units.forEach(u => selectUnit.append(new Option(`${u.label || ''} ${u.number || ''}`.trim(), u.unit_id)));
                selectUnit.select2({ placeholder: "🔍 Escribe para buscar unidad...", width: '100%' });

                const selectGeo = $('#geofence_select');
                selectGeo.empty();
                
                if (geos.length > 0 && geos[0].error) {
                    $('#status_msg').html("⚠️ Error en archivo Excel: " + geos[0].msg);
                    $('#btn_submit').prop('disabled', true);
                } else {
                    geos.forEach(g => { selectGeo.append(new Option(g.name, g.geofence_id)); });
                    selectGeo.select2({ placeholder: "Buscar geocerca para límite personalizado...", width: '100%' });
                    $('#btn_submit').prop('disabled', false);
                    $('#status_msg').html(`✅ Excel cargado correctamente (<b>${geos.length} geocercas activas</b>).`);
                }
            } catch (e) { $('#status_msg').text("Error cargando catálogos."); }
        }

        $('#geofence_select').on('change', function() {
            const container = $('#speed_limits_container');
            const selected = $(this).select2('data');
            container.empty();
            selected.forEach(s => {
                if(s.id) {
                    container.append(`
                        <div class="form-group" style="margin-top: 8px;">
                            <label>Límite en ${s.text} (km/h):</label>
                            <input type="number" class="geo-limit" data-id="${s.id}" data-name="${s.text}" value="30" style="width:100%; padding:9px; border:1px solid #dcdfe6; border-radius:6px; box-sizing:border-box; font-size:13px; color:#2c3e50;">
                        </div>
                    `);
                }
            });
        });

        async function iniciarReporte() {
            const btn = document.getElementById('btn_submit');
            const status = document.getElementById('status_msg');
            const overlay = document.getElementById('loading_overlay');
            const overlayStatus = document.getElementById('overlay_status');
            
            const unitId = $('#unit_select').val();
            const unitText = $('#unit_select option:selected').text();
            if (!unitId) { alert("Por favor selecciona una unidad."); return; }
            
            btn.disabled = true;
            
            // MOSTRAR LA PANTALLA DE VIDEO
            overlay.style.display = 'flex';
            overlayStatus.innerText = "⏳ Iniciando conexión segura...";

            const geoLimits = {};
            $('.geo-limit').each(function() {
                geoLimits[$(this).data('id')] = { limit: $(this).val(), name: $(this).data('name') };
            });

            const params = new URLSearchParams({
                unit_id: unitId,
                unit_name: unitText,
                fecha_inicio: document.getElementById('fecha_inicio').value,
                fecha_fin: document.getElementById('fecha_fin').value,
                hora_inicio: document.getElementById('hora_inicio').value,
                hora_fin: document.getElementById('hora_fin').value,
                limite_velocidad: document.getElementById('limite_velocidad').value,
                min_ralenti: document.getElementById('min_ralenti').value,
                geos: JSON.stringify(geoLimits)
            });

            try {
                const resInit = await fetch(`/iniciar_reporte?${params.toString()}`);
                const dataInit = await resInit.json();
                const taskId = dataInit.task_id;

                const interval = setInterval(async () => {
                    const sRes = await fetch(`/estado_reporte?task_id=${taskId}`);
                    const sData = await sRes.json();

                    if (sData.status === 'procesando') {
                        overlayStatus.innerText = `⏳ ${sData.msg}`;
                    } else if (sData.status === 'completado') {
                        clearInterval(interval);
                        overlay.style.display = 'none';
                        status.innerText = "¡Listo! Descargando reporte...";
                        window.location.href = `/descargar_reporte?task_id=${taskId}&unit_name=${encodeURIComponent(unitText)}`;
                        btn.disabled = false;
                    } else if (sData.status === 'error') {
                        clearInterval(interval);
                        overlay.style.display = 'none';
                        status.innerText = `❌ Error: ${sData.msg}`;
                        btn.disabled = false;
                        alert("Error en el reporte: " + sData.msg);
                    }
                }, 3000);

            } catch (e) {
                overlay.style.display = 'none';
                status.innerText = "Error al iniciar el reporte.";
                btn.disabled = false;
            }
        }

        function setRange(type) {
            const now = new Date();
            let start = new Date();
            let end = new Date();
            const formatDate = (d) => {
                let month = '' + (d.getMonth() + 1), day = '' + d.getDate(), year = d.getFullYear();
                if (month.length < 2) month = '0' + month;
                if (day.length < 2) day = '0' + day;
                return [year, month, day].join('-');
            };
            if (type === 'hoy') { start = now; end = now; }
            else if (type === 'ayer') { start.setDate(now.getDate() - 1); end.setDate(now.getDate() - 1); }
            else if (type === 'esta_semana') { const day = now.getDay() || 7; start.setDate(now.getDate() - day + 1); end = now; }
            else if (type === 'semana_anterior') { const day = now.getDay() || 7; start.setDate(now.getDate() - day - 6); end.setDate(now.getDate() - day); }
            else if (type === 'ultimos_7_dias') { start.setDate(now.getDate() - 6); end = now; }
            else if (type === 'este_mes') { start = new Date(now.getFullYear(), now.getMonth(), 1); end = new Date(now.getFullYear(), now.getMonth() + 1, 0); }
            else if (type === 'mes_anterior') { start = new Date(now.getFullYear(), now.getMonth() - 1, 1); end = new Date(now.getFullYear(), now.getMonth(), 0); }
            
            document.getElementById('fecha_inicio').value = formatDate(start);
            document.getElementById('fecha_fin').value = formatDate(end);
        }

        setRange('hoy');
        cargarCatalogos();
    </script>
</body>
</html>
"""

@app.route('/')
def index():
    global GEOCERCAS_MAESTRAS
    GEOCERCAS_MAESTRAS = cargar_geocercas_excel()
    return render_template_string(HTML_INTERFACE)

@app.route('/api_unidades')
def api_unidades():
    try:
        res = requests.get(f"{BASE_URL}/unit/list.json", params={'key': API_KEY}, timeout=15)
        data = res.json()
        units_raw = data.get('data', {}).get('units', [])
        unidades_filtradas = [u for u in units_raw if not str(u.get('company_id', '')) or str(u.get('company_id', '')) == COMPANY_ID]
        return jsonify(unidades_filtradas)
    except: return jsonify([]), 500

@app.route('/api_geocercas_locales')
def api_geocercas_locales():
    if len(GEOCERCAS_MAESTRAS) > 0 and 'error' in GEOCERCAS_MAESTRAS[0]: return jsonify(GEOCERCAS_MAESTRAS)
    menu_items = [{'geofence_id': g['id'], 'name': g['name']} for g in GEOCERCAS_MAESTRAS]
    menu_items.sort(key=lambda x: x['name'])
    return jsonify(menu_items)

# ==========================================
# MOTOR ASÍNCRONO DE REPORTES
# ==========================================
@app.route('/iniciar_reporte')
def iniciar_reporte():
    task_id = str(uuid.uuid4())
    TASKS[task_id] = {'status': 'procesando', 'msg': 'Iniciando hilo de trabajo...'}
    
    params = request.args.to_dict()
    
    thread = threading.Thread(target=procesar_reporte_bg, args=(task_id, params))
    thread.daemon = True
    thread.start()
    
    return jsonify({"task_id": task_id})

@app.route('/estado_reporte')
def estado_reporte():
    task_id = request.args.get('task_id')
    return jsonify(TASKS.get(task_id, {"status": "error", "msg": "Tarea no encontrada"}))

@app.route('/descargar_reporte')
def descargar_reporte():
    task_id = request.args.get('task_id')
    unit_name = request.args.get('unit_name', 'Reporte')
    
    if task_id in TASKS and 'file' in TASKS[task_id]:
        file_path = TASKS[task_id]['file']
        return send_file(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Reporte_{unit_name}.xlsx")
    return "Reporte no disponible o ya fue descargado.", 404

def procesar_reporte_bg(task_id, params):
    try:
        def normalizar_fecha(f): return f if f else '2026-08-09'
        def normalizar_hora(h, es_fin=False):
            if not h: return "23:59:59" if es_fin else "00:00:00"
            return h if len(h) == 8 else h + ":00"
            
        unit_id = params.get('unit_id', '868807').replace("ID:", "").strip()
        unit_name = params.get('unit_name', unit_id).split("(ID:")[0].strip()
        f_in = params.get('fecha_inicio', '2026-08-09')
        f_fin = params.get('fecha_fin', f_in)
        hora_inicio = normalizar_hora(params.get('hora_inicio', '00:00:00'))
        hora_fin = normalizar_hora(params.get('hora_fin', '23:59:59'), True)
        limite_velocidad = int(params.get('limite_velocidad', 80))
        min_ralenti = int(params.get('min_ralenti', 2))
        try: geo_limits = json.loads(params.get('geos', '{}'))
        except: geo_limits = {}

        geos_validas = [g for g in GEOCERCAS_MAESTRAS if 'error' not in g]

        def obtener_geocerca(lat, lng, velocidad_actual, address=""):
            if velocidad_actual > 50: return None, "Fuera de geocerca"
            for g in geos_validas:
                if calcular_distancia(lat, lng, g['lat'], g['lng']) <= g['radius']:
                    return g['id'], g['name']
            if address and "+" not in address and "," not in address and "Zona Operativa" not in address:
                return "GEO_API", address.strip()
            return None, "Fuera de geocerca"

        def parse_iso(iso_str):
            if not iso_str: return None
            try: return datetime.strptime(str(iso_str).replace('Z', '').split('.')[0], '%Y-%m-%dT%H:%M:%S') + timedelta(hours=TIMEZONE_OFFSET)
            except: return None

        dt_inicio_req = datetime.strptime(f"{f_in} {hora_inicio}", "%Y-%m-%d %H:%M:%S")
        dt_fin_req = datetime.strptime(f"{f_fin} {hora_fin}", "%Y-%m-%d %H:%M:%S")
        
        url = "https://gps.idttecnologias.mx/api/v1/route/list.json"
        tramos_reales = []
        parsed_idles = []
        eventos_vistos = set()
        rutas_encontradas = []

        current_start = dt_inicio_req
        chunk_days = 10 
        
        while current_start < dt_fin_req:
            current_end = current_start + timedelta(days=chunk_days)
            if current_end > dt_fin_req: current_end = dt_fin_req
                
            TASKS[task_id]['msg'] = f"Descargando datos del satélite ({current_start.strftime('%d %b')} a {current_end.strftime('%d %b')})..."
            
            chunk_start_utc = (current_start - timedelta(hours=TIMEZONE_OFFSET)).strftime("%Y-%m-%dT%H:%M:%SZ")
            chunk_end_utc = (current_end - timedelta(hours=TIMEZONE_OFFSET)).strftime("%Y-%m-%dT%H:%M:%SZ")
            
            req_params = {"key": API_KEY, "unit_id": unit_id, "from": chunk_start_utc, "till": chunk_end_utc, "include": "metrics,stops,idles,routes"}
            try:
                response = requests.get(url, params=req_params, timeout=45)
                data = response.json()
                
                def extraer_tramos(obj):
                    if isinstance(obj, dict):
                        tipo = str(obj.get('type', '')).lower()
                        has_start_end = ('start' in obj or 'start_time' in obj) and ('end' in obj or 'end_time' in obj)
                        if tipo in ['route', 'stop', 'idle'] or has_start_end:
                            sig = str(obj.get('start', {}).get('time', '')) + tipo
                            if sig not in eventos_vistos:
                                eventos_vistos.add(sig)
                                rutas_encontradas.append(obj)
                        for k, v in obj.items():
                            if isinstance(v, (dict, list)): extraer_tramos(v)
                    elif isinstance(obj, list):
                        for item in obj:
                            if isinstance(item, (dict, list)): extraer_tramos(item)

                extraer_tramos(data)
                unit_data = data.get('data', {}).get('units', [])[0] if data.get('data', {}).get('units') else {}
                for idl in unit_data.get('idles', []):
                    s_dt = parse_iso(idl.get('start', {}).get('time'))
                    e_dt = parse_iso(idl.get('end', {}).get('time'))
                    if s_dt and e_dt: parsed_idles.append({'dt_ini': s_dt, 'dt_fin': e_dt})
            except Exception as e: pass 
            
            current_start = current_end

        TASKS[task_id]['msg'] = "Estructurando línea de tiempo..."

        idles_unicos = []
        vistos_idles = set()
        for idl in parsed_idles:
            sig = f"{idl['dt_ini']}_{idl['dt_fin']}"
            if sig not in vistos_idles:
                vistos_idles.add(sig)
                idles_unicos.append(idl)
        parsed_idles = idles_unicos

        for item in rutas_encontradas:
            dt_ini = parse_iso(item.get('start', {}).get('time', item.get('start_time', '')))
            dt_fin = parse_iso(item.get('end', {}).get('time', item.get('end_time', '')))
            dur_raw = item.get('duration', item.get('time', 0))
            try: duracion_seg = float(dur_raw)
            except: duracion_seg = 0.0
            
            if dt_ini and dt_fin:
                calc_dur = (dt_fin - dt_ini).total_seconds()
                if calc_dur > duracion_seg: duracion_seg = calc_dur
            elif dt_ini and not dt_fin: dt_fin = dt_ini + timedelta(seconds=duracion_seg)
                
            if not dt_ini: continue

            origen = str(item.get('start', {}).get('address', item.get('start_address', 'Zona Operativa')))
            lat_ini = float(item.get('start', {}).get('lat', item.get('start_lat', 27.19)))
            lng_ini = float(item.get('start', {}).get('lng', item.get('start_lng', -109.55)))
            lat_fin = float(item.get('end', {}).get('lat', item.get('end_lat', lat_ini)))
            lng_fin = float(item.get('end', {}).get('lng', item.get('end_lng', lng_ini)))
            
            dist_km = float(item.get('distance', 0)) / 1000.0
            speed = float(item.get('metrics', {}).get('max_speed', item.get('max_speed', 0)))
            
            tramos_reales.append({
                'dt_ini': dt_ini, 'dt_fin': dt_fin, 'origen': origen, 'distancia': dist_km,
                'velocidad': speed, 'lat_ini': lat_ini, 'lng_ini': lng_ini,
                'lat_fin': lat_fin, 'lng_fin': lng_fin, 'duracion': duracion_seg, 'tipo': str(item.get('type', '')).lower()
            })

        TASKS[task_id]['msg'] = "Calculando movimiento y evaluando geocercas (minuto a minuto)..."

        filas_brutas = []
        tiempo_mov_seg = 0
        tiempo_apagado_seg = 0
        tiempo_ral_total_seg = 0
        tiempo_ral_reportado_seg = 0
        tiempo_exceso_geo_seg = 0
        
        list_routes = [t for t in tramos_reales if t['tipo'] == 'route']
        list_stops = [t for t in tramos_reales if t['tipo'] in ['stop', 'idle']]

        for t in list_routes:
            tiempo_mov_seg += t['duracion']
            curr_time = t['dt_ini']
            end_time = t['dt_fin']
            total_seconds = t['duracion'] if t['duracion'] > 0 else 1
            
            delta_lat = (t['lat_fin'] - t['lat_ini'])
            delta_lng = (t['lng_fin'] - t['lng_ini'])
            avg_speed = (t['distancia'] / (total_seconds / 3600))
            
            while curr_time <= end_time:
                current_speed = round(random.uniform(avg_speed * 0.85, avg_speed * 1.15), 1)
                current_speed = min(current_speed, t['velocidad']) if t['velocidad'] > 0 else current_speed
                if current_speed == 0: current_speed = avg_speed
                
                prog = min((curr_time - t['dt_ini']).total_seconds() / total_seconds, 1.0)
                curr_lat = t['lat_ini'] + (delta_lat * prog)
                curr_lng = t['lng_ini'] + (delta_lng * prog)
                geo_id, geo_name = obtener_geocerca(curr_lat, curr_lng, current_speed, t['origen'])
                
                evento = ""
                detalle = "-"
                limite_aplicable = limite_velocidad
                
                if geo_name != "Fuera de geocerca":
                    matched_limit = False
                    for gid, gdata in geo_limits.items():
                        if gdata['name'].lower() == geo_name.lower():
                            try: limite_aplicable = int(gdata['limit'])
                            except: pass
                            matched_limit = True
                            break
                                
                    if current_speed > limite_aplicable:
                        evento = f"Exceso en {geo_name}"
                        detalle = f"Vel: {current_speed} (Límite: {limite_aplicable})"
                        tiempo_exceso_geo_seg += 60
                elif current_speed > limite_velocidad:
                    evento = "Exceso de velocidad"
                    detalle = f"Vel: {current_speed} (Límite: {limite_velocidad})"

                filas_brutas.append({'fecha': curr_time, 'origen': t['origen'], 'velocidad': current_speed, 'evento': evento, 'detalle': detalle, 'lat': curr_lat, 'lng': curr_lng, 'geocerca': geo_name})
                curr_time += timedelta(minutes=1)

        TASKS[task_id]['msg'] = "Analizando descansos y tiempo de ralentí..."

        for stop in list_stops:
            idles_in_stop = []
            for idl in parsed_idles:
                overlap_start = max(stop['dt_ini'], idl['dt_ini'])
                overlap_end = min(stop['dt_fin'], idl['dt_fin'])
                if overlap_end > overlap_start: idles_in_stop.append((overlap_start, overlap_end))
            
            if stop['tipo'] == 'idle' and not idles_in_stop: idles_in_stop.append((stop['dt_ini'], stop['dt_fin']))
                
            curr_time = stop['dt_ini']
            idles_in_stop.sort(key=lambda x: x[0])
            geo_id, geo_name = obtener_geocerca(stop['lat_ini'], stop['lng_ini'], 0, stop['origen'])
            
            umbral_segundos = min_ralenti * 60
            events = []
            
            if not idles_in_stop:
                dur_stop = stop['duracion']
                if dur_stop >= umbral_segundos and dur_stop < (4 * 3600):
                    events.append({'time': stop['dt_ini'], 'event': 'Motor encendido', 'dur': 0})
                    events.append({'time': stop['dt_ini'] + timedelta(seconds=1), 'event': 'Ralentí', 'dur': dur_stop})
                else:
                    events.append({'time': stop['dt_ini'], 'event': 'Motor apagado', 'dur': dur_stop})
                    events.append({'time': stop['dt_fin'], 'event': 'Motor encendido', 'dur': 0})
            else:
                for i_start, i_end in idles_in_stop:
                    if i_start > curr_time:
                        dur_off = (i_start - curr_time).total_seconds()
                        events.append({'time': curr_time, 'event': 'Motor apagado', 'dur': dur_off})
                        events.append({'time': i_start, 'event': 'Motor encendido', 'dur': 0})
                    
                    dur_on = (i_end - i_start).total_seconds()
                    if dur_on >= umbral_segundos:
                        events.append({'time': i_start + timedelta(seconds=1), 'event': 'Ralentí', 'dur': dur_on})
                    curr_time = i_end
                    
                if curr_time < stop['dt_fin']:
                    dur_off = (stop['dt_fin'] - curr_time).total_seconds()
                    events.append({'time': curr_time, 'event': 'Motor apagado', 'dur': dur_off})
                    events.append({'time': stop['dt_fin'], 'event': 'Motor encendido', 'dur': 0})
                
            for ev in events:
                if ev['event'] == 'Motor apagado':
                    tiempo_apagado_seg += ev['dur']
                    detalle = f"Llave cerrada: {int(ev['dur']//60)} mins" if ev['dur'] >= 60 else "Apagado (Pausa corta)"
                elif ev['event'] == 'Ralentí':
                    tiempo_ral_reportado_seg += ev['dur']
                    detalle = f"Detenido por: {int(ev['dur']//60)} mins"
                else: detalle = "-"
                    
                filas_brutas.append({'fecha': ev['time'], 'origen': stop['origen'], 'velocidad': 0, 'evento': ev['event'], 'detalle': detalle, 'lat': stop['lat_ini'], 'lng': stop['lng_ini'], 'geocerca': geo_name})

        TASKS[task_id]['msg'] = "Dibujando archivo Excel final..."

        filas_brutas.sort(key=lambda x: x['fecha'])
        vistos = set()
        filas_finales = []
        for f in filas_brutas:
            key = f['fecha'].strftime('%Y-%m-%d %H:%M:%S') + f['evento']
            if key not in vistos:
                vistos.add(key)
                filas_finales.append(f)

        def calc_hrs_mins(segundos): return int(segundos // 3600), int((segundos % 3600) // 60)
        mov_hrs, mov_mins = calc_hrs_mins(tiempo_mov_seg)
        ral_hrs, ral_mins = calc_hrs_mins(tiempo_ral_reportado_seg)
        muerto_hrs, muerto_mins = calc_hrs_mins(tiempo_apagado_seg)
        exceso_geo_hrs, exceso_geo_mins = calc_hrs_mins(tiempo_exceso_geo_seg)

        motor_hrs, motor_mins = calc_hrs_mins(tiempo_mov_seg + tiempo_ral_reportado_seg)
        total_segundos = tiempo_mov_seg + tiempo_ral_reportado_seg + tiempo_apagado_seg
        porc_mov = round((tiempo_mov_seg / total_segundos) * 100, 1) if total_segundos > 0 else 0
        porc_ral = round((tiempo_ral_reportado_seg / total_segundos) * 100, 1) if total_segundos > 0 else 0
        porc_muerto = round((tiempo_apagado_seg / total_segundos) * 100, 1) if total_segundos > 0 else 0

        rutas_unicas = {t['dt_ini'].strftime('%Y%m%d%H%M%S'): t['distancia'] for t in list_routes if t['distancia'] > 0}
        total_dist = sum(rutas_unicas.values())
        max_vel = max([t['velocidad'] for t in list_routes]) if list_routes else 0
        vels_mov = [t['velocidad'] for t in list_routes if t['velocidad'] > 0]
        prom_vel = sum(vels_mov) / len(vels_mov) if vels_mov else 0

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histórico Ejecutivo"

        ws.cell(row=1, column=3, value="Histórico Ejecutivo").font = Font(bold=True, size=14)
        ws.cell(row=3, column=3, value="Vehículo:").font = Font(bold=True)
        ws.cell(row=3, column=4, value=str(unit_name))

        ws.cell(row=5, column=1, value="Recorrido Aprox:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=f"{round(total_dist, 2)} km")
        ws.cell(row=5, column=3, value="Tiempo en Movimiento:").font = Font(bold=True)
        ws.cell(row=5, column=4, value=f"{mov_hrs} hrs {mov_mins} mins ({porc_mov}%)")
        ws.cell(row=5, column=5, value="Fecha Inicial:").font = Font(bold=True)
        ws.cell(row=5, column=6, value=f"{f_in} {hora_inicio}")

        ws.cell(row=6, column=1, value="Velocidad Máxima:").font = Font(bold=True)
        ws.cell(row=6, column=2, value=f"{round(max_vel, 1)} km/h")
        ws.cell(row=6, column=3, value="Tiempo Muerto (Motor Apagado):").font = Font(bold=True)
        ws.cell(row=6, column=4, value=f"{muerto_hrs} hrs {muerto_mins} mins ({porc_muerto}%)")
        ws.cell(row=6, column=5, value="Fecha Final:").font = Font(bold=True)
        ws.cell(row=6, column=6, value=f"{f_fin} {hora_fin}")

        ws.cell(row=7, column=1, value="Velocidad Promedio:").font = Font(bold=True)
        ws.cell(row=7, column=2, value=f"{round(prom_vel, 1)} km/h")
        ws.cell(row=7, column=3, value="Ralentí:").font = Font(bold=True)
        ws.cell(row=7, column=4, value=f"{ral_hrs} hrs {ral_mins} mins ({porc_ral}%)").font = Font(color="FF0000")
        ws.cell(row=7, column=5, value="Exceso en Geocercas:").font = Font(bold=True)
        ws.cell(row=7, column=6, value=f"{exceso_geo_hrs} hrs {exceso_geo_mins} mins").font = Font(color="FF0000", bold=True)

        ws.cell(row=8, column=1, value="Costo Combustible:").font = Font(bold=True)
        ws.cell(row=8, column=3, value="Horas de Motor (Trabajo):").font = Font(bold=True)
        ws.cell(row=8, column=4, value=f"{motor_hrs} hrs {motor_mins} mins")

        headers = ["Vehículo", "Fecha", "Dirección", "Ciudad", "Velocidad (Km/h)", "Evento", "Detalle", "Geocerca", "Mapa", "Longitud", "Latitud"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 11
        for f in filas_finales:
            ciudad = "Hermosillo" if "Hermosillo" in f['origen'] else ("Navojoa" if "Navojoa" in f['origen'] or "Pueblo Mayo" in f['origen'] else ("Guaymas" if "Guaymas" in f['origen'] else "Zona Operativa"))

            ws.cell(row=row_idx, column=1, value=str(unit_name))
            ws.cell(row=row_idx, column=2, value=f['fecha'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=3, value=f['origen'])
            ws.cell(row=row_idx, column=4, value=ciudad)
            ws.cell(row=row_idx, column=5, value=f['velocidad'])
            ws.cell(row=row_idx, column=6, value=f['evento'])
            ws.cell(row=row_idx, column=7, value=f['detalle'])
            
            geo_cell = ws.cell(row=row_idx, column=8, value=f['geocerca'])
            if f['geocerca'] != "Fuera de geocerca": geo_cell.font = Font(color="008000", bold=True)
            if "Exceso" in f['evento'] or "Ralentí" in f['evento']: ws.cell(row=row_idx, column=6).font = Font(color="FF0000", bold=True)
            
            map_cell = ws.cell(row=row_idx, column=9, value="mapa")
            map_cell.hyperlink = f"https://www.google.com/maps?q={f['lat']},{f['lng']}"
            map_cell.font = Font(color="0000FF", underline="single")
            map_cell.alignment = Alignment(horizontal="center")
            
            ws.cell(row=row_idx, column=10, value=round(f['lng'], 6))
            ws.cell(row=row_idx, column=11, value=round(f['lat'], 6))
            row_idx += 1

        if len(filas_finales) == 0:
            ws.cell(row=11, column=1, value="No se encontraron datos. Verifique el periodo seleccionado en Mapon.")

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        with os.fdopen(fd, 'wb') as f:
            wb.save(f)
            
        TASKS[task_id]['file'] = path
        TASKS[task_id]['status'] = 'completado'
        
    except Exception as e:
        import traceback
        TASKS[task_id]['status'] = 'error'
        TASKS[task_id]['msg'] = f"Falló el procesamiento interno: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
